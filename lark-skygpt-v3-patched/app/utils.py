import io
import logging
import mimetypes
from typing import Optional, Tuple, List
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from .config import settings

logger = logging.getLogger(__name__)
TZ = ZoneInfo(settings.TIMEZONE)

# Optional deps
try:
    from pypdf import PdfReader
    HAVE_PYPDF = True
except Exception:
    HAVE_PYPDF = False

try:
    import docx
    HAVE_DOCX = True
except Exception:
    HAVE_DOCX = False

try:
    import openpyxl
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

def now_local() -> datetime:
    return datetime.now(TZ)

def yesterday_range_local() -> Tuple[datetime, datetime]:
    today = now_local().date()
    y = today - timedelta(days=1)
    start = datetime(y.year, y.month, y.day, 0, 0, 0, tzinfo=TZ)
    end = start + timedelta(days=1)
    return start, end

def to_epoch_ms(dt: datetime) -> int:
    return int(dt.timestamp() * 1000)

def guess_filename(default_name: str, content_type: Optional[str], header_name: Optional[str]) -> str:
    if header_name:
        return header_name
    if content_type:
        ext = mimetypes.guess_extension(content_type.split(";")[0].strip()) or ""
        if ext and not default_name.endswith(ext):
            return default_name + ext
    return default_name

def safe_decode_text(data: bytes) -> str:
    for enc in ("utf-8", "utf-16", "big5", "gbk", "latin1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")

def extract_text_from_pdf(data: bytes) -> str:
    if not HAVE_PYPDF:
        return "[PDF 已接收，但伺服器未安裝 pypdf]"
    try:
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages[:20]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                pass
        text = "\n".join(parts).strip()
        return text or "[PDF 無可抽取文字]"
    except Exception:
        return "[PDF 解析失敗]"

def extract_text_from_docx(data: bytes) -> str:
    if not HAVE_DOCX:
        return "[Word 已接收，但伺服器未安裝 python-docx]"
    try:
        document = docx.Document(io.BytesIO(data))
        text = "\n".join([p.text for p in document.paragraphs])
        return text.strip() or "[Word 文件無可抽取文字]"
    except Exception:
        return "[Word 文件解析失敗]"

def extract_text_from_excel(data: bytes, max_sheets: int = 5, max_rows: int = 100, max_cols: int = 30) -> str:
    if not HAVE_OPENPYXL:
        return "[Excel 已接收，但伺服器未安裝 openpyxl]"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        out = []
        for name in wb.sheetnames[:max_sheets]:
            sh = wb[name]
            out.append(f"--- 工作表: {name} ---")
            for r in sh.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
                row = ["" if v is None else str(v) for v in r]
                if any(cell != "" for cell in row):
                    out.append("\t".join(row))
        text = "\n".join(out).strip()
        if not text:
            return "[Excel 檔案無內容]"
        if len(wb.sheetnames) > max_sheets or sh.max_row > max_rows or sh.max_column > max_cols:
            text += f"\n...（可能已截斷，僅列前 {max_sheets} 張表、{max_rows} 行、{max_cols} 欄）"
        return text
    except Exception:
        return "[Excel 解析失敗]"

def extract_text_generic(data: bytes, filename: str, content_type: Optional[str]) -> str:
    name = filename.lower()
    ct = (content_type or "").lower()

    if name.endswith(".pdf") or "pdf" in ct:
        return extract_text_from_pdf(data)
    if name.endswith(".docx") or "officedocument.wordprocessingml.document" in ct:
        return extract_text_from_docx(data)
    if name.endswith((".xlsx", ".xlsm")) or "officedocument.spreadsheetml.sheet" in ct:
        return extract_text_from_excel(data)
    if name.endswith(".csv") or "csv" in ct:
        try:
            return "\n".join(safe_decode_text(data).splitlines()[:200])
        except Exception:
            return "[CSV 解析失敗]"
    if name.endswith(".txt") or "text/plain" in ct:
        return safe_decode_text(data)

    try:
        return safe_decode_text(data)
    except Exception:
        return f"[{filename}（{len(data)} bytes）]"
