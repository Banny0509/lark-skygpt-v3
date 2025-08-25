import os
import json
import time
import asyncio
import logging
from typing import Any, Dict

import httpx
from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, PlainTextResponse

from .config import settings
from .openai_client import (
    reply_text_or_fallback,
    summarize_text_or_fallback,
    describe_image_from_message_or_fallback,
    describe_pdf_from_message_or_fallback,
    _download_message_resource,  # 從訊息端點下載檔案
)

# 若有 tasks 就寫庫供每日摘要用；沒有也不影響主流程
try:
    from . import tasks  # noqa
    _HAS_TASKS = True
except Exception:
    tasks = None  # type: ignore
    _HAS_TASKS = False

logger = logging.getLogger("sky_lark")
logging.basicConfig(level=logging.INFO)

app = FastAPI()

# =========================
# Lark Token / 發送文字（輕量 helper）
# =========================
LARK_TENANT_TOKEN_URL = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
LARK_SEND_MESSAGE_URL = "https://open.larksuite.com/open-apis/im/v1/messages"

def _resolve_lark_credentials() -> tuple[str, str]:
    app_id = (
        getattr(settings, "LARK_APP_ID", None)
        or getattr(settings, "FEISHU_APP_ID", None)
        or getattr(settings, "APP_ID", None)
        or os.getenv("LARK_APP_ID")
        or os.getenv("FEISHU_APP_ID")
        or os.getenv("APP_ID")
    )
    app_secret = (
        getattr(settings, "LARK_APP_SECRET", None)
        or getattr(settings, "FEISHU_APP_SECRET", None)
        or getattr(settings, "APP_SECRET", None)
        or os.getenv("LARK_APP_SECRET")
        or os.getenv("FEISHU_APP_SECRET")
        or os.getenv("APP_SECRET")
    )
    if not app_id or not app_secret:
        raise RuntimeError("缺少 Lark 憑證：請設定 LARK_APP_ID/LARK_APP_SECRET（或 FEISHU_/APP_ 對應）")
    return app_id, app_secret

async def _get_tenant_access_token(http: httpx.AsyncClient) -> str:
    app_id, app_secret = _resolve_lark_credentials()
    r = await http.post(LARK_TENANT_TOKEN_URL, json={"app_id": app_id, "app_secret": app_secret}, timeout=20)
    r.raise_for_status()
    tok = r.json().get("tenant_access_token")
    if not tok:
        raise RuntimeError("取得 tenant_access_token 失敗")
    return tok

async def reply_text(http: httpx.AsyncClient, chat_id: str, text: str, *, by_chat_id: bool = True) -> None:
    try:
        token = await _get_tenant_access_token(http)
    except Exception as e:
        logger.error("無法發送訊息：%s", e)
        return
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    params = {"receive_id_type": "chat_id" if by_chat_id else "open_id"}
    body = {"receive_id": chat_id, "msg_type": "text", "content": json.dumps({"text": text}, ensure_ascii=False)}
    try:
        resp = await http.post(LARK_SEND_MESSAGE_URL, headers=headers, params=params, json=body, timeout=20)
        if resp.status_code >= 400:
            errtxt = (await resp.aread()).decode(errors="ignore")
            logger.error("回覆訊息失敗 %s: %s", resp.status_code, errtxt)
    except Exception as e:
        logger.exception("回覆訊息發送異常：%s", e)

# =========================
# 健康檢查
# =========================
@app.get("/")
async def root_ok():
    return {"ok": True, "service": "web", "env": getattr(settings, "ENV", "prod")}

@app.get("/healthz")
async def healthz():
    return {"ok": True, "env": getattr(settings, "ENV", "prod")}

@app.get("/health")
async def health_alias():
    return await healthz()

# =========================
# 去重（Lark 會重試）
# =========================
try:
    import redis.asyncio as aioredis
except Exception:
    aioredis = None

REDIS_URL = getattr(settings, "REDIS_URL", None) or os.getenv("REDIS_URL") or os.getenv("REDIS_CONNECTION_URL")
_redis = aioredis.from_url(REDIS_URL) if (aioredis and REDIS_URL) else None
_local_seen: dict[str, float] = {}
_LOCAL_TTL = 24 * 3600

async def _dedupe_mark(message_id: str) -> bool:
    key = f"lark:msg:{message_id}"
    if _redis:
        try:
            ok = await _redis.setnx(key, "1")
            if ok:
                await _redis.expire(key, _LOCAL_TTL)
                return True
            return False
        except Exception:
            pass
    now = time.time()
    for k, ts in list(_local_seen.items()):
        if now - ts > _LOCAL_TTL:
            _local_seen.pop(k, None)
    if key in _local_seen:
        return False
    _local_seen[key] = now
    return True

# =========================
# 判斷群組是否 @ 機器人
# =========================
def _is_p2p_chat(msg: dict) -> bool:
    chat_type = (msg.get("chat_type") or "").lower()
    return chat_type in ("p2p", "single", "private", "p2p_chat")

def _normalize_mention_token(x: Any) -> str:
    try:
        if x is None:
            return ""
        if isinstance(x, str):
            return x.lower()
        if isinstance(x, dict):
            return json.dumps(x, ensure_ascii=False, sort_keys=True).lower()
        return str(x).lower()
    except Exception:
        return ""

def _bot_is_mentioned(msg: dict, content_text: str) -> bool:
    mentions = msg.get("mentions") or []
    app_id = (
        getattr(settings, "LARK_APP_ID", None)
        or getattr(settings, "FEISHU_APP_ID", None)
        or getattr(settings, "APP_ID", None)
        or os.getenv("LARK_APP_ID")
        or os.getenv("FEISHU_APP_ID")
        or os.getenv("APP_ID")
        or ""
    ).lower()
    for m in mentions:
        mid = m.get("id") if isinstance(m, dict) else m
        token = _normalize_mention_token(mid)
        if app_id and app_id in token:
            return True
        if isinstance(mid, str) and app_id and app_id in mid.lower():
            return True
    text = content_text or ""
    if "<at" in text:  # 富文本
        return True
    bot_name = os.getenv("BOT_NAME") or ""
    if bot_name and (f"@{bot_name}" in text):
        return True
    if "@bot" in text:
        return True
    return False

# =========================
# Webhook：固定 /webhook/lark（即時回 200、背景處理）
# =========================
@app.post("/webhook/lark")
async def lark_webhook(request: Request):
    try:
        event = await request.json()
    except Exception:
        return JSONResponse({"msg": "ok"}, status_code=200)

    if isinstance(event, dict) and "challenge" in event:
        return JSONResponse({"challenge": event["challenge"]}, status_code=200)

    asyncio.create_task(_process_lark_event(event))
    return JSONResponse({"msg": "ok"}, status_code=200)

# 相容舊路由
@app.post("/lark/webhook")
async def lark_webhook_alias(request: Request):
    return await lark_webhook(request)

# =========================
# 背景處理：僅群組被 @ 才回；私聊照舊
# =========================
async def _process_lark_event(event: Dict[str, Any]) -> None:
    header = event.get("header", {}) or {}
    if header.get("event_type") != "im.message.receive_v1":
        return

    ev = event.get("event", {}) or {}
    msg = ev.get("message", {}) or {}
    msg_type = msg.get("message_type")
    message_id = msg.get("message_id")
    chat_id = msg.get("chat_id")
    content_raw = msg.get("content", "{}")

    # 非阻塞寫庫（若有）
    if _HAS_TASKS and hasattr(tasks, "record_message"):
        try:
            asyncio.create_task(tasks.record_message(event))
        except Exception as _e:
            logger.debug("record_message 啟動失敗：%s", _e)

    try:
        content = json.loads(content_raw) if isinstance(content_raw, str) else (content_raw or {})
    except Exception:
        content = {}

    if not chat_id:
        return
    if message_id:
        first = await _dedupe_mark(message_id)
        if not first:
            return

    text_in = (content.get("text") or "") if isinstance(content, dict) else ""
    require_mention = not _is_p2p_chat(msg)
    mentioned = _bot_is_mentioned(msg, text_in)

    async with httpx.AsyncClient() as http:
        # 文字
        if msg_type == "text":
            if require_mention and not mentioned:
                return
            text = (text_in or "").strip()
            if not text:
                return
            try:
                if text in ("摘要", "總結", "总结", "summary"):
                    reply = await summarize_text_or_fallback(http, text)
                else:
                    reply = await reply_text_or_fallback(http, text)
            except Exception as e:
                logger.exception("處理文字訊息失敗：%s", e)
                reply = "(降級) 處理文字訊息時發生例外，已記錄日誌。"
            await reply_text(http, chat_id, reply, by_chat_id=True)
            return

        # 圖片
        if msg_type == "image":
            if require_mention and not mentioned:
                return
            image_key = content.get("image_key")
            if not (message_id and image_key):
                await reply_text(http, chat_id, "(降級) 缺少 message_id 或 image_key。", by_chat_id=True)
                return
            try:
                result = await describe_image_from_message_or_fallback(http, message_id, image_key)
            except Exception as e:
                logger.exception("圖片 Vision 解析失敗：%s", e)
                result = f"(降級) 圖像解析異常：{e}"
            await reply_text(http, chat_id, result, by_chat_id=True)
            return

        # 檔案：PDF（@ 才處理）；DOCX/XLSX 可選擇開啟
        if msg_type == "file":
            if require_mention and not mentioned:
                return
            file_key = content.get("file_key")
            file_name = (content.get("file_name") or "").lower()
            if not (message_id and file_key):
                await reply_text(http, chat_id, "(降級) 缺少 message_id 或 file_key。", by_chat_id=True)
                return

            if file_name.endswith(".pdf"):
                try:
                    result = await describe_pdf_from_message_or_fallback(http, message_id, file_key)
                except Exception as e:
                    logger.exception("PDF Vision 解析失敗：%s", e)
                    result = f"(降級) PDF 解析異常：{e}"
                await reply_text(http, chat_id, result, by_chat_id=True)
                return

            # 其他格式提示（若要啟用 DOCX/XLSX 摘要，打開下方兩段）
            # if file_name.endswith(".docx"):
            #     data = await _download_message_resource(http, message_id, file_key, rtype="file")
            #     text = _extract_docx_text(data)
            #     result = await summarize_text_or_fallback(http, text)
            #     await reply_text(http, chat_id, result, by_chat_id=True); return
            # if file_name.endswith(".xlsx"):
            #     data = await _download_message_resource(http, message_id, file_key, rtype="file")
            #     text = _extract_xlsx_text(data)
            #     result = await summarize_text_or_fallback(http, text)
            #     await reply_text(http, chat_id, result, by_chat_id=True); return

            await reply_text(
                http, chat_id,
                f"(提示) 已接收檔案：{file_name or file_key}。目前僅對 PDF 走 Vision；如需 DOCX/XLSX 請告知。",
                by_chat_id=True,
            )
            return

        # 其他型別：略過
        return

# ===== DOCX / XLSX 輔助（若要啟用） =====
def _extract_docx_text(data: bytes) -> str:
    try:
        from io import BytesIO
        from docx import Document
        doc = Document(BytesIO(data))
        parts = [p.text for p in doc.paragraphs if p.text]
        return "\n".join(parts).strip() or "(空白文件)"
    except Exception as e:
        return f"(降級) DOCX 解析失敗：{e}"

def _extract_xlsx_text(data: bytes, max_rows: int = 50, max_cols: int = 20) -> str:
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > max_rows:
                rows.append("...（已截斷）"); break
            cells = []
            for c_idx, v in enumerate(row, start=1):
                if c_idx > max_cols:
                    cells.append("…"); break
                cells.append("" if v is None else str(v))
            rows.append("\t".join(cells))
        return f"[工作表：{ws.title}]\n" + "\n".join(rows)
    except Exception as e:
        return f"(降級) XLSX 解析失敗：{e}"
